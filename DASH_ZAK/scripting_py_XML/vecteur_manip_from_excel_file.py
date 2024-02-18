 #!/usr/bin/python3
 import openpyxl
 
 
 wb = openpyxl.load_workbook('/Users/zakaria/Downloads/medaf- Copie2.xlsx')
 print(wb.sheetnames)
 #print("\ncellule A1")
 #wb1 = wb.sheetnames
 for i in wb.sheetnames:
     v = 1
     w = wb[i]
     c = w['A1'].value
     print(c)
     #for v in w.iter_rows(values_only=True):
         #print(v)
     while v < 10:
         print(w.cell(row=v, column=1).value)
         v += 1
 print('salina')
